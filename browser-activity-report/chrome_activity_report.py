#!/usr/bin/env python3
"""
chrome_activity_report.py

Builds a categorized browsing-activity workbook from Chrome "History" SQLite files.

  - One worksheet per user (multiple workstation captures are merged and de-duplicated)
  - Category + Flag columns driven by an editable rule table (see CATEGORY_RULES)
  - Summary, Downloads, Sources (with SHA-256), and Uncategorized Domains sheets

Usage:
    python chrome_activity_report.py --input "C:\Reports\Captures"
    python chrome_activity_report.py -i . --outdir out --start 08-01-2026

Requires: openpyxl   (pip install openpyxl)
Optional: tzdata on Windows for local-time conversion (pip install tzdata)
"""

import argparse
import hashlib
import os
import re
import shutil
import sqlite3
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------

# Set to your local IANA timezone, e.g. "America/New_York", "Europe/London".
LOCAL_TZ_NAME = "America/New_York"

# Display format for date cells. These are real datetime values, so this only
# controls how Excel renders them -- sorting and Date Filters are unaffected.
#   Excel codes: mm-dd-yyyy hh:mm:ss  |  yyyy-mm-dd hh:mm:ss  |  m/d/yy h:mm AM/PM
EXCEL_DATE_FORMAT = "mm-dd-yyyy hh:mm:ss"
PY_DATE_FMT = "%m-%d-%Y"          # headings and Sources sheet
PY_DATETIME_FMT = "%m-%d-%Y %H:%M"

# Pin any file whose name the parser gets wrong (e.g. three-part usernames).
#   "ChromeHistory_first.last.middleELMRI01": ("first.last.middle", "ELMRI01"),
FILENAME_OVERRIDES = {}

# Work-hours window used only to populate the "On Clock" column.
WORK_DAYS = {0, 1, 2, 3, 4}          # Mon-Fri
WORK_START = (8, 0)                   # 08:00
WORK_END = (17, 0)                    # 17:00

# Chrome transition core types. AUTO_SUBFRAME is embedded/ad content the user
# never clicked, so it is dropped by default.
TRANSITION_CORE = {
    0: "Link", 1: "Typed", 2: "Bookmark", 3: "Auto Subframe",
    4: "Manual Subframe", 5: "Generated", 6: "Start Page", 7: "Form Submit",
    8: "Reload", 9: "Keyword", 10: "Keyword Generated",
}
DROP_CORE_TYPES = {3}                 # Auto Subframe
QUAL_CLIENT_REDIRECT = 0x01000000
QUAL_SERVER_REDIRECT = 0x80000000

# --------------------------------------------------------------------------
# CATEGORY RULES  -- first match wins, so order matters.
#
#   (Category, Flag, [patterns])
#
# A pattern beginning with "path:" is matched against the full lowercased URL.
# Everything else is matched against the hostname.
# Flag: "" | "Review" | "High"
# --------------------------------------------------------------------------

# === CATEGORY RULES =========================================================
#
# First match wins, so ORDER MATTERS. A pattern beginning with "path:" matches
# against the full lowercased URL; anything else matches against the hostname.
#
# Flags:  "High"   -> red, worth a closer look
#         "Review" -> amber, personal but commonly tolerated
#         ""       -> neutral / work-related
#
# The sections marked CUSTOMIZE are placeholders. Add your own organization's
# domains there (your intranet, your line-of-business apps, your vendors), then
# re-run. The "Uncategorized Domains" sheet in the output shows you which
# domains had no rule so you know what to add.

CATEGORY_RULES = [

    # ---- CUSTOMIZE: your own company / internal systems --------------------
    # Put your intranet, SSO, and line-of-business hostnames here so normal
    # work traffic is labelled instead of landing in "Uncategorized".
    # ("Company / Internal", "", [
    #     "yourcompany.com", "intranet", ".local", ".lan",
    # ]),

    # ---- CUSTOMIZE: your industry / line-of-business apps ------------------
    # e.g. an EHR, an ERP, a CRM, a case-management system.
    # ("Line of Business", "", [
    #     "yourapp.example.com",
    # ]),

    # --- Job searching -- checked FIRST so an employer's own careers page on
    # --- a shared ATS host doesn't get mislabelled. Add your ATS tenant to the
    # --- "Company / Internal" block above if you want internal postings split
    # --- out from external ones.
    ("Job Searching", "High", [
        "indeed.com", "ziprecruiter.com", "glassdoor.com", "monster.com",
        "careerbuilder.com", "simplyhired.com", "snagajob.com", "dice.com",
        "usajobs.gov", "joinhandshake.com", "flexjobs.com", "jobs.net",
        "myworkdayjobs.com", "workday.com", "icims.com", "greenhouse.io",
        "lever.co", "taleo.net", "jobvite.com", "smartrecruiters.com",
        "ashbyhq.com", "resume.io", "zippia.com", "livecareer.com",
        "path:linkedin.com/jobs", "path:/careers", "path:careers.",
        "path:jobs.", "path:/job-openings", "path:/apply",
    ]),

    # --- Education / career development ------------------------------------
    ("Education / Career Dev", "Review", [
        "coursera.org", "udemy.com", "edx.org", "khanacademy.org",
        "wgu.edu", "snhu.edu", "phoenix.edu", "linkedin.com/learning",
    ]),

    # --- Microsoft 365 / general work tooling -----------------------------
    ("Work Tools / M365", "", [
        "outlook.office.com", "outlook.office365.com", "outlook.cloud.microsoft",
        "teams.microsoft.com", "teams.cloud.microsoft", "word.cloud.microsoft",
        "excel.cloud.microsoft", "m365.cloud.microsoft", "office.com",
        "microsoft365.com", "login.microsoftonline.com", "microsoftonline.com",
        "aka.ms", "support.microsoft.com", "onedrive.live.com",
        "sharepoint.com", "attachments.office.net", "adobe.com",
        "acrobat.adobe.com", "docusign.net", "docusign.com", "zoom.us",
        "webex.com", "gotomeeting.com", "ringcentral.com", "slack.com",
    ]),

    # --- CUSTOMIZE: sanctioned vendor / partner portals --------------------
    # Portals your staff legitimately use (suppliers, clearinghouses, payers,
    # partner systems). Listing them keeps work traffic out of "Uncategorized".
    # ("Vendor / Partner Portal", "", [
    #     "vendor1.example.com", "partner2.example.net",
    # ]),

    # --- FLAGGED: adult ---------------------------------------------------
    ("Adult Content", "High", [
        "pornhub.com", "xvideos.com", "xnxx.com", "redtube.com",
        "onlyfans.com", "xhamster.com", "chaturbate.com", "stripchat.com",
        "adultfriendfinder.com",
    ]),

    # --- FLAGGED: gambling ------------------------------------------------
    ("Gambling", "High", [
        "draftkings.com", "fanduel.com", "betmgm.com", "caesars.com",
        "bovada.lv", "stake.com", "betonline.ag", "chumbacasino.com",
        "goldennugget.com",
    ]),

    # --- FLAGGED: personal cloud storage / transfer (exfiltration risk) ----
    ("Personal File Transfer", "High", [
        "wetransfer.com", "mega.nz", "sendspace.com", "file.io",
        "dropbox.com", "box.com", "pcloud.com", "mediafire.com",
        "smash.com", "filemail.com", "transfernow.net",
    ]),

    # --- FLAGGED: generative AI (data-disclosure risk) --------------------
    ("AI / Chatbot", "Review", [
        "chatgpt.com", "chat.openai.com", "openai.com", "claude.ai",
        "gemini.google.com", "bard.google.com", "perplexity.ai",
        "copilot.microsoft.com", "poe.com", "character.ai", "deepseek.com",
    ]),

    # --- FLAGGED: personal webmail ----------------------------------------
    ("Personal Email", "Review", [
        "mail.google.com", "mail.yahoo.com", "outlook.live.com",
        "mail.com", "aol.com", "proton.me", "protonmail.com",
        "zoho.com", "icloud.com", "gmx.com", "yandex.com",
    ]),

    # --- FLAGGED: social media --------------------------------------------
    ("Social Media", "Review", [
        "facebook.com", "fb.com", "messenger.com", "instagram.com",
        "twitter.com", "x.com", "tiktok.com", "snapchat.com", "reddit.com",
        "pinterest.com", "linkedin.com", "tumblr.com", "threads.net",
        "discord.com", "nextdoor.com", "whatsapp.com", "vk.com",
        "bereal.com", "twitch.tv",
    ]),

    # --- FLAGGED: shopping ------------------------------------------------
    ("Shopping", "Review", [
        "amazon.com", "ebay.com", "walmart.com", "target.com", "etsy.com",
        "temu.com", "shein.com", "wish.com", "aliexpress.com", "wayfair.com",
        "homedepot.com", "lowes.com", "bestbuy.com", "costco.com",
        "samsclub.com", "kohls.com", "poshmark.com", "mercari.com",
        "shop.app", "chewy.com", "ulta.com", "sephora.com",
    ]),

    # --- FLAGGED: streaming / entertainment -------------------------------
    ("Streaming / Entertainment", "Review", [
        "netflix.com", "hulu.com", "disneyplus.com", "max.com", "hbomax.com",
        "peacocktv.com", "paramountplus.com", "primevideo.com",
        "spotify.com", "pandora.com", "iheart.com", "soundcloud.com",
        "crunchyroll.com", "roku.com", "tubitv.com", "plex.tv",
    ]),
    ("Video / YouTube", "Review", [
        "youtube.com", "youtu.be", "vimeo.com", "dailymotion.com",
    ]),

    # --- FLAGGED: gaming --------------------------------------------------
    ("Gaming", "Review", [
        "steampowered.com", "steamcommunity.com", "epicgames.com",
        "roblox.com", "miniclip.com", "coolmathgames.com", "poki.com",
        "solitaired.com", "chess.com", "king.com",
        "path:nytimes.com/games",
    ]),

    # --- FLAGGED: personal banking / finance ------------------------------
    ("Banking / Personal Finance", "Review", [
        "chase.com", "wellsfargo.com", "bankofamerica.com", "usbank.com",
        "capitalone.com", "discover.com", "citi.com", "pnc.com",
        "usaa.com", "navyfederal.org", "paypal.com", "venmo.com",
        "cash.app", "creditkarma.com", "irs.gov", "ssa.gov",
        "turbotax.intuit.com",
    ]),

    # --- Neutral / informational ------------------------------------------
    ("News / Weather", "", [
        "cnn.com", "foxnews.com", "nbcnews.com", "abcnews.go.com",
        "cbsnews.com", "npr.org", "apnews.com", "reuters.com",
        "usatoday.com", "nytimes.com", "washingtonpost.com",
        "weather.com", "weather.gov", "accuweather.com", "wunderground.com",
    ]),
    ("Search Engine", "", [
        "path:google.com/search", "path:bing.com/search", "duckduckgo.com",
        "search.yahoo.com", "path:search.brave.com", "ecosia.org",
    ]),
    ("Reference / Encyclopedia", "", [
        "wikipedia.org", "britannica.com", "stackoverflow.com",
        "wolframalpha.com",
    ]),
    ("Travel / Maps", "", [
        "path:google.com/maps", "mapquest.com", "waze.com",
        "expedia.com", "booking.com", "airbnb.com", "tripadvisor.com",
    ]),
    ("Food / Restaurants", "Review", [
        "doordash.com", "ubereats.com", "grubhub.com", "yelp.com",
        "toasttab.com", "olo.com", "chick-fil-a.com", "mcdonalds.com",
        "starbucks.com", "dominos.com", "papajohns.com",
    ]),

    # --- Noise to suppress -------------------------------------------------
    ("Ad / Tracker / CDN", "", [
        "doubleclick.net", "googlesyndication.com", "googleadservices.com",
        "google-analytics.com", "googletagmanager.com", "scorecardresearch.com",
        "adservice.google", "adnxs.com", "criteo.com", "taboola.com",
        "outbrain.com", "cloudfront.net", "akamaihd.net", "gstatic.com",
        "googleapis.com", "cloudflare.com", "licdn.com", "cdninstagram.com",
        "bat.bing.com", "hotjar.com", "segment.io", "amazon-adsystem.com",
        "storage.googleapis.com", "onecdn.static.microsoft",
    ]),
    ("Browser / System / Local", "", [
        "chrome", "edge", "about:blank", "localhost", "127.0.0.1",
        "file", "blob", "accounts.google.com", "clients2.google.com",
        "safebrowsing", "update.googleapis.com", "ntp.msn.com",
    ]),
]

NOISE_CATEGORIES = {"Ad / Tracker / CDN", "Browser / System / Local"}

# --- Search-term overrides ------------------------------------------------
# Applied only when a visit landed on a search engine. A search-engine hit is
# uninformative on its own; the typed query is what shows intent.
#   (regex, Category, Flag)
TERM_RULES = [
    (r"\b(cover letter|resume|resum[e\u00e9]|cv template|job application|"
     r"jobs? near me|hiring|now hiring|apply for|application for employment|"
     r"interview questions|salary for|pay rate for|indeed|ziprecruiter)\b",
     "Job Searching", "High"),
    (r"\b(meme|memes|funny|joke|jokes|trivia)\b",
     "Personal Browsing (Search)", "Review"),
    (r"\b(recipe|restaurant|menu|near me food|takeout|bakery)\b",
     "Food / Restaurants", "Review"),
    (r"\b(degree|certificate|college|tuition|classes|certification|"
     r"licensure)\b",
     "Education / Career Dev", "Review"),
]


FLAG_FILLS = {
    "High": PatternFill("solid", fgColor="FFC7CE"),
    "Review": PatternFill("solid", fgColor="FFEB9C"),
}
FLAG_FONTS = {
    "High": Font(name="Arial", size=10, color="9C0006"),
    "Review": Font(name="Arial", size=10, color="9C6500"),
}

HEADER_FILL = PatternFill("solid", fgColor="0B5C7C")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=13, bold=True, color="0B5C7C")
THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CHROME_EPOCH = datetime(1601, 1, 1, tzinfo=timezone.utc)

# --------------------------------------------------------------------------
# HELPERS
# --------------------------------------------------------------------------


def get_local_tz():
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo(LOCAL_TZ_NAME), LOCAL_TZ_NAME
    except Exception:
        print(f"  ! Could not load timezone '{LOCAL_TZ_NAME}'. "
              f"Times will be UTC. On Windows run: pip install tzdata")
        return timezone.utc, "UTC"


def chrome_time(value, tz):
    """Chrome stores microseconds since 1601-01-01 UTC."""
    if not value:
        return None
    try:
        return (CHROME_EPOCH + timedelta(microseconds=int(value))).astimezone(tz)
    except (OverflowError, OSError, ValueError):
        return None


def parse_filename(stem):
    """
    'ChromeHistory_first.lastPC01' -> ('first.last', 'PC01')

    The username is first.last (lowercase); the workstation is appended with no
    separator and starts with a capital letter. Anything that does not parse
    cleanly can be pinned in FILENAME_OVERRIDES below.
    """
    if stem in FILENAME_OVERRIDES:
        return FILENAME_OVERRIDES[stem]

    # Strip an upload/export prefix such as "1787605854539_".
    stem = re.sub(r"^\d{8,}[_-]", "", stem)
    # Some transfers replace the dot in first.last with an underscore.
    stem = re.sub(r"^(ChromeHistory[_-]?[A-Za-z]+)_([a-z]+)", r"\1.\2", stem)

    # Strip the browser prefix first, case-insensitively. The main match must
    # stay case-SENSITIVE: the capital letter starting the computer name is the
    # only thing marking where the username ends.
    browser = ""
    bm = re.match(r"^(Chrome|Edge|Brave|Browser)History[_-]?", stem, re.I)
    if bm:
        kind = bm.group(1).lower()
        if kind not in ("chrome", "browser"):
            browser = f" ({bm.group(1).title()})"
        stem = stem[bm.end():]

    m = re.match(r"^(?P<user>[A-Za-z]+\.[a-z]+?)(?P<host>[A-Z][A-Za-z0-9_.-]*)$", stem)
    if m:
        return m.group("user").lower(), m.group("host") + browser

    m = re.match(r"^(?P<user>[A-Za-z]+\.[A-Za-z]+)$", stem)
    if m:
        return m.group("user").lower(), "Unknown" + browser
    return stem.lower(), "Unknown" + browser


def hostname_of(url):
    u = (url or "").strip()
    low = u.lower()
    for scheme in ("chrome://", "chrome-extension://", "edge://", "about:", "file://"):
        if low.startswith(scheme):
            return scheme.rstrip(":/")
    m = re.match(r"^[a-z][a-z0-9+.-]*://([^/?#]+)", low)
    host = m.group(1) if m else low.split("/")[0]
    host = host.split("@")[-1].split(":")[0]
    return host[4:] if host.startswith("www.") else host


def registrable(host):
    """Cheap eTLD+1. Good enough for reporting, not for adjudication."""
    if not host or host.replace(".", "").isdigit() or "://" in host:
        return host
    parts = host.split(".")
    if len(parts) <= 2:
        return host
    two_part = {"co.uk", "org.uk", "ac.uk", "gov.uk", "com.au", "co.nz",
                "co.jp", "com.br", "co.in", "ms.gov", "state.ms.us"}
    if ".".join(parts[-2:]) in two_part and len(parts) >= 3:
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


SEARCH_CATEGORIES = {"Search Engine"}


def categorize(url, host, term=""):
    low_url = (url or "").lower()
    low_host = (host or "").lower()
    category, flag = "Uncategorized", "Review"
    for cat, fl, patterns in CATEGORY_RULES:
        matched = False
        for pat in patterns:
            if pat.startswith("path:"):
                if pat[5:] in low_url:
                    matched = True
            elif pat in low_host:
                matched = True
            if matched:
                break
        if matched:
            category, flag = cat, fl
            break

    # A search-engine hit says nothing on its own; the typed term does.
    if term and category in SEARCH_CATEGORIES:
        low_term = term.lower()
        for pattern, cat, fl in TERM_RULES:
            if re.search(pattern, low_term):
                return cat, fl
    return category, flag


def sha256_of(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def on_clock(dt):
    if dt is None:
        return ""
    if dt.weekday() not in WORK_DAYS:
        return "No"
    start = dt.replace(hour=WORK_START[0], minute=WORK_START[1], second=0, microsecond=0)
    end = dt.replace(hour=WORK_END[0], minute=WORK_END[1], second=0, microsecond=0)
    return "Yes" if start <= dt <= end else "No"


# --------------------------------------------------------------------------
# EXTRACTION
# --------------------------------------------------------------------------


def read_history(path, user, host, tz, keep_subframes=False, keep_redirects=False):
    """Copy the DB to temp (avoids locks / read-only shares) and extract visits."""
    visits, downloads = [], []
    tmpdir = tempfile.mkdtemp(prefix="chromehist_")
    tmp = os.path.join(tmpdir, "History")
    try:
        shutil.copy2(path, tmp)
        for suffix in ("-wal", "-shm"):
            side = path + suffix
            if os.path.exists(side):
                shutil.copy2(side, tmp + suffix)

        con = sqlite3.connect(tmp)
        con.text_factory = lambda b: b.decode("utf-8", "replace")
        cur = con.cursor()

        terms = {}
        try:
            cur.execute("SELECT url_id, term FROM keyword_search_terms")
            terms = {row[0]: row[1] for row in cur.fetchall()}
        except sqlite3.Error:
            pass

        cur.execute("""
            SELECT u.id, u.url, u.title, v.visit_time, v.visit_duration, v.transition
            FROM visits v JOIN urls u ON u.id = v.url
        """)
        for url_id, url, title, vtime, vdur, transition in cur.fetchall():
            t = int(transition or 0) & 0xFFFFFFFF
            core = t & 0xFF
            if not keep_subframes and core in DROP_CORE_TYPES:
                continue
            is_redirect = bool(t & (QUAL_CLIENT_REDIRECT | QUAL_SERVER_REDIRECT))
            if not keep_redirects and is_redirect:
                continue

            hostname = hostname_of(url)
            domain = registrable(hostname)
            term = terms.get(url_id, "")
            category, flag = categorize(url, hostname, term)
            dt = chrome_time(vtime, tz)
            visits.append({
                "dt": dt,
                "user": user,
                "host": host,
                "category": category,
                "flag": flag,
                "domain": domain,
                "title": (title or "").strip()[:300],
                "url": (url or "")[:1000],
                "term": term,
                "transition": TRANSITION_CORE.get(core, f"Type {core}"),
                "duration": round((vdur or 0) / 1_000_000, 1),
                "on_clock": on_clock(dt),
            })

        try:
            cur.execute("""
                SELECT target_path, tab_url, start_time, received_bytes, mime_type
                FROM downloads
            """)
            for tpath, taburl, stime, nbytes, mime in cur.fetchall():
                downloads.append({
                    "dt": chrome_time(stime, tz),
                    "user": user,
                    "host": host,
                    "file": os.path.basename(tpath or ""),
                    "path": tpath or "",
                    "source": (taburl or "")[:500],
                    "domain": registrable(hostname_of(taburl)),
                    "size_mb": round((nbytes or 0) / 1_048_576, 2),
                    "mime": mime or "",
                })
        except sqlite3.Error:
            pass

        con.close()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    return visits, downloads


# --------------------------------------------------------------------------
# WORKBOOK
# --------------------------------------------------------------------------

VISIT_COLUMNS = [
    ("Date / Time", 19), ("On Clock", 9), ("Workstation", 15),
    ("Category", 26), ("Flag", 9), ("Domain", 28),
    ("Page Title", 52), ("Search Term", 28), ("How Reached", 14),
    ("Seconds", 9), ("Full URL", 70),
]


def style_header(ws, row, columns):
    for idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 20


def write_user_sheet(wb, user, rows, tz_label):
    ws = wb.create_sheet(title=user[:31])
    ws.cell(row=1, column=1, value=f"Browsing activity - {user}").font = TITLE_FONT
    span = ""
    dated = [r["dt"] for r in rows if r["dt"]]
    if dated:
        span = (f"{min(dated).strftime(PY_DATE_FMT)} to "
                f"{max(dated).strftime(PY_DATE_FMT)}  ")
    ws.cell(row=2, column=1,
            value=f"{span}{len(rows):,} page visits.  All times {tz_label}. "
                  f"Subframe and redirect noise removed.").font = Font(
        name="Arial", size=9, italic=True, color="595959")

    style_header(ws, 4, VISIT_COLUMNS)

    for i, r in enumerate(sorted(rows, key=lambda x: (x["dt"] is None, x["dt"])), start=5):
        values = [
            r["dt"].replace(tzinfo=None, microsecond=0) if r["dt"] else "",
            r["on_clock"], r["host"], r["category"], r["flag"], r["domain"],
            r["title"], r["term"], r["transition"], r["duration"], r["url"],
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = FLAG_FONTS.get(r["flag"], BODY_FONT)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c in (7, 8)))
            if c == 1 and r["dt"]:
                cell.number_format = EXCEL_DATE_FORMAT
            if r["flag"] in FLAG_FILLS:
                cell.fill = FLAG_FILLS[r["flag"]]

    last = max(5, 4 + len(rows))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(VISIT_COLUMNS))}{last}"
    ws.freeze_panes = "A5"
    return ws


def write_summary(wb, per_user, tz_label, generated):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.cell(row=1, column=1, value="Browsing Activity Report").font = Font(
        name="Arial", size=16, bold=True, color="0B5C7C")
    ws.cell(row=2, column=1,
            value=f"Generated {generated.strftime(PY_DATETIME_FMT)} ({tz_label}). "
                  f"Counts exclude ads, trackers, and browser system pages.").font = Font(
        name="Arial", size=9, italic=True, color="595959")

    categories = sorted({r["category"] for rows in per_user.values() for r in rows}
                        - NOISE_CATEGORIES)
    users = sorted(per_user)

    header_row = 4
    cols = [("Category", 30), ("Flag", 10)] + [(u, 18) for u in users] + [("Total", 12)]
    style_header(ws, header_row, cols)

    flag_of = {}
    for rows in per_user.values():
        for r in rows:
            flag_of.setdefault(r["category"], r["flag"])

    counts = {u: Counter(r["category"] for r in per_user[u]) for u in users}
    order = {"High": 0, "Review": 1, "": 2}
    categories.sort(key=lambda c: (order.get(flag_of.get(c, ""), 2), c))

    row = header_row + 1
    for cat in categories:
        flag = flag_of.get(cat, "")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=flag)
        for j, u in enumerate(users, start=3):
            ws.cell(row=row, column=j, value=counts[u].get(cat, 0))
        first = get_column_letter(3)
        lastc = get_column_letter(2 + len(users))
        ws.cell(row=row, column=3 + len(users),
                value=f"=SUM({first}{row}:{lastc}{row})")
        for c in range(1, 4 + len(users)):
            cell = ws.cell(row=row, column=c)
            cell.font = FLAG_FONTS.get(flag, BODY_FONT)
            cell.border = CELL_BORDER
            if flag in FLAG_FILLS:
                cell.fill = FLAG_FILLS[flag]
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="Total visits").font = Font(
        name="Arial", size=10, bold=True)
    for c in range(1, 4 + len(users)):
        ws.cell(row=total_row, column=c).border = CELL_BORDER
        ws.cell(row=total_row, column=c).font = Font(name="Arial", size=10, bold=True)
    for j in range(3, 4 + len(users)):
        col = get_column_letter(j)
        ws.cell(row=total_row, column=j,
                value=f"=SUM({col}{header_row + 1}:{col}{total_row - 1})")

    note = total_row + 2
    ws.cell(row=note, column=1, value="How to read this report").font = Font(
        name="Arial", size=11, bold=True, color="0B5C7C")
    for i, line in enumerate([
        "Red rows are categories worth a closer look. Amber rows are personal but "
        "commonly tolerated use.",
        "A flag is a starting point for a conversation, not a finding. Check the "
        "'On Clock' column before drawing conclusions.",
        "Category assignment is automated by domain matching. Review the "
        "'Uncategorized Domains' sheet and re-run after tuning the rules.",
        "One visit is one page load. A single task can generate many rows.",
    ], start=1):
        ws.cell(row=note + i, column=1, value=line).font = Font(
            name="Arial", size=9, color="404040")
    ws.freeze_panes = f"A{header_row + 1}"
    return ws


def write_sources(wb, sources):
    ws = wb.create_sheet(title="Sources")
    ws.cell(row=1, column=1, value="Source files").font = TITLE_FONT
    cols = [("File", 44), ("User", 18), ("Workstation", 16), ("Size (MB)", 11),
            ("Visits Kept", 12), ("First Visit", 19), ("Last Visit", 19),
            ("SHA-256", 66)]
    style_header(ws, 3, cols)
    for i, s in enumerate(sources, start=4):
        for c, v in enumerate([s["file"], s["user"], s["host"], s["size_mb"],
                               s["count"], s["first"], s["last"], s["sha256"]], start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
    ws.freeze_panes = "A4"


def write_downloads(wb, downloads, tz_label):
    ws = wb.create_sheet(title="Downloads")
    ws.cell(row=1, column=1, value="Downloaded files").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"All times {tz_label}. Worth reviewing for anything containing "
                  f"patient data.").font = Font(name="Arial", size=9, italic=True,
                                                color="595959")
    cols = [("Date / Time", 19), ("User", 18), ("Workstation", 15), ("File Name", 40),
            ("Source Site", 26), ("Size (MB)", 10), ("Type", 22), ("Saved To", 56)]
    style_header(ws, 4, cols)
    rows = sorted(downloads, key=lambda d: (d["dt"] is None, d["dt"]))
    for i, d in enumerate(rows, start=5):
        for c, v in enumerate([
            d["dt"].replace(tzinfo=None, microsecond=0) if d["dt"] else "",
            d["user"], d["host"], d["file"], d["domain"], d["size_mb"],
            d["mime"], d["path"]], start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            if c == 1 and d["dt"]:
                cell.number_format = EXCEL_DATE_FORMAT
    if rows:
        ws.auto_filter.ref = f"A4:H{4 + len(rows)}"
    ws.freeze_panes = "A5"


def write_uncategorized(wb, per_user):
    ws = wb.create_sheet(title="Uncategorized Domains")
    ws.cell(row=1, column=1, value="Domains no rule matched").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Add these to CATEGORY_RULES in the script and re-run to tighten "
                  "the report.").font = Font(name="Arial", size=9, italic=True,
                                             color="595959")
    tally = Counter()
    example = {}
    for rows in per_user.values():
        for r in rows:
            if r["category"] == "Uncategorized":
                tally[r["domain"]] += 1
                example.setdefault(r["domain"], r["title"] or r["url"])
    style_header(ws, 4, [("Domain", 36), ("Visits", 10), ("Example Page", 76)])
    for i, (dom, n) in enumerate(tally.most_common(), start=5):
        for c, v in enumerate([dom, n, example.get(dom, "")], start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
    ws.freeze_panes = "A5"


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------


def parse_bound(text):
    """Accept MM-DD-YYYY or YYYY-MM-DD; the 4-digit part disambiguates."""
    for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"Could not read date '{text}'. Use MM-DD-YYYY or YYYY-MM-DD.")


def discover_files(folder):
    """Return [(path, user, workstation), ...] for capture files in a folder."""
    out = []
    for name in sorted(os.listdir(folder)):
        full = os.path.join(folder, name)
        if not os.path.isfile(full) or name.endswith(("-wal", "-shm")):
            continue
        if "history" not in name.lower():
            continue
        out.append((full,) + parse_filename(strip_ext(os.path.basename(full))))
    return out


def strip_ext(name):
    """These files usually have no extension, and the username contains a dot."""
    for ext in (".sqlite", ".db", ".sqlite3", ".bak"):
        if name.lower().endswith(ext):
            return name[: -len(ext)]
    return name


def build_reports(entries, outdir, start=None, end=None, keep_noise=False,
                  keep_subframes=False, keep_redirects=False, progress=None):
    """
    entries : [(path, user, workstation), ...]
    start/end : tz-aware datetimes or None
    progress : optional callable(message, fraction_0_to_1)

    Returns (written_paths, stats_by_user).
    """
    def say(msg, frac=None):
        if progress:
            progress(msg, frac)
        else:
            print(msg)

    tz, tz_label = get_local_tz()
    per_user = defaultdict(list)
    all_downloads, sources = [], []
    seen = set()
    total = max(1, len(entries))

    for i, (full, user, host) in enumerate(entries):
        say(f"Reading {os.path.basename(full)}  ->  {user} @ {host}", i / total)
        try:
            visits, downloads = read_history(full, user, host, tz,
                                             keep_subframes, keep_redirects)
        except sqlite3.DatabaseError as exc:
            say(f"   SKIPPED, not a readable SQLite database ({exc})")
            continue

        kept = []
        for v in visits:
            if v["dt"]:
                if start and v["dt"] < start:
                    continue
                if end and v["dt"] > end:
                    continue
            if not keep_noise and v["category"] in NOISE_CATEGORIES:
                continue
            key = (user, v["url"], v["dt"].isoformat() if v["dt"] else "")
            if key in seen:
                continue
            seen.add(key)
            kept.append(v)

        per_user[user].extend(kept)
        for d in downloads:
            if d["dt"]:
                if start and d["dt"] < start:
                    continue
                if end and d["dt"] > end:
                    continue
            all_downloads.append(d)

        dated = [v["dt"] for v in kept if v["dt"]]
        sources.append({
            "file": os.path.basename(full),
            "user": user,
            "host": host,
            "size_mb": round(os.path.getsize(full) / 1_048_576, 2),
            "count": len(kept),
            "first": min(dated).strftime(PY_DATETIME_FMT) if dated else "",
            "last": max(dated).strftime(PY_DATETIME_FMT) if dated else "",
            "sha256": sha256_of(full),
        })
        say(f"   {len(kept):,} visits kept "
            f"({len(visits) - len(kept):,} dropped as noise/duplicate/out-of-range)")

    generated = datetime.now(tz)
    os.makedirs(outdir, exist_ok=True)
    written, stats = [], {}

    for user in sorted(per_user):
        rows = per_user[user]
        one = {user: rows}
        wb = Workbook()
        wb.remove(wb.active)
        write_summary(wb, one, tz_label, generated)
        write_user_sheet(wb, user, rows, tz_label)
        user_dl = [d for d in all_downloads if d["user"] == user]
        write_downloads(wb, user_dl, tz_label)
        write_uncategorized(wb, one)
        write_sources(wb, [x for x in sources if x["user"] == user])

        safe = re.sub(r"[^A-Za-z0-9._-]", "_", user)
        path = os.path.join(outdir, f"Browsing_Activity_{safe}.xlsx")
        wb.save(path)
        written.append(path)
        stats[user] = {
            "visits": len(rows),
            "high": sum(1 for r in rows if r["flag"] == "High"),
            "downloads": len(user_dl),
        }
        say(f"Wrote {path}")

    say("Done.", 1.0)
    return written, stats


def main():
    ap = argparse.ArgumentParser(description="Chrome history -> categorized XLSX report")
    ap.add_argument("-i", "--input", default=".", help="folder holding the ChromeHistory_* files")
    ap.add_argument("-o", "--outdir", default=".",
                    help="folder to write one workbook per user into")
    ap.add_argument("--start", help="only include visits on/after this date (MM-DD-YYYY)")
    ap.add_argument("--end", help="only include visits on/before this date (MM-DD-YYYY)")
    ap.add_argument("--keep-noise", action="store_true",
                    help="keep ad/tracker/system rows in the user sheets")
    ap.add_argument("--keep-subframes", action="store_true",
                    help="keep auto-loaded subframe hits (very noisy)")
    ap.add_argument("--keep-redirects", action="store_true",
                    help="keep intermediate redirect hops")
    args = ap.parse_args()

    tz, _ = get_local_tz()
    entries = discover_files(args.input)
    if not entries:
        sys.exit(f"No history files found in {args.input}")

    start = parse_bound(args.start).replace(tzinfo=tz) if args.start else None
    end = (parse_bound(args.end).replace(hour=23, minute=59, second=59, tzinfo=tz)
           if args.end else None)

    written, stats = build_reports(
        entries, args.outdir, start, end,
        args.keep_noise, args.keep_subframes, args.keep_redirects)

    print()
    for path in written:
        user = os.path.basename(path)[len("Browsing_Activity_"):-len(".xlsx")]
        st = stats.get(user, {})
        print(f"  {path}")
        print(f"      {st.get('visits', 0):,} visits, {st.get('high', 0):,} high-flag, "
              f"{st.get('downloads', 0):,} downloads")


if __name__ == "__main__":
    main()
